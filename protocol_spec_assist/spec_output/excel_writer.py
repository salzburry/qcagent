"""
Excel workbook writer — formatted .xlsx output matching the 9-tab program spec template.

Tab layout:
  1.Cover | 2.QC Review | 3.Data Prep | 4.StudyPop |
  5A.Demos | 5B.ClinChars | 5C.BioVars | 5D.LabVars |
  6.TreatVars | 7.Outcomes

Variable tabs share a standardised column layout:
  Time Period | Variable | Label | Values | Definition |
  Code Lists Group | Additional Notes | Date Modified | QC Reviewed

Requires openpyxl.
"""

from __future__ import annotations
from pathlib import Path

from .spec_schema import ProgramSpec

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Styling constants ──────────────────────────────────────────────────────

if OPENPYXL_AVAILABLE:
    FILL_HEADER = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    FILL_SECTION = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
    FILL_DATA = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    FILL_NOTE = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    FILL_GREEN = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    FILL_YELLOW = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    FILL_RED = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
    FONT_SECTION = Font(color="FFFFFF", bold=True, size=11)
    FONT_NORMAL = Font(size=10)
    FONT_BOLD = Font(size=10, bold=True)
    ALIGN_WRAP = Alignment(vertical="top", wrap_text=True)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
else:
    FILL_HEADER = FILL_SECTION = FILL_DATA = FILL_NOTE = None
    FILL_GREEN = FILL_YELLOW = FILL_RED = None
    FONT_HEADER = FONT_SECTION = FONT_NORMAL = FONT_BOLD = None
    ALIGN_WRAP = ALIGN_CENTER = None


def _confidence_fill(conf: float | None):
    if conf is None:
        return None
    if conf >= 0.8:
        return FILL_GREEN
    if conf >= 0.6:
        return FILL_YELLOW
    return FILL_RED


def _write_header_row(ws, row: int, headers: list[str]):
    """Write a dark-red header row matching the sample spec."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER


def _write_section_row(ws, row: int, title: str, ncols: int):
    """Write a green section heading row spanning all columns."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = ALIGN_CENTER
    for col in range(2, ncols + 1):
        ws.cell(row=row, column=col).fill = FILL_SECTION


def _write_data_row(ws, row: int, values: list, confidence: float | None = None):
    """Write a data row with optional confidence-based fill."""
    fill = _confidence_fill(confidence) if confidence else FILL_DATA
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_WRAP
        if fill:
            cell.fill = fill


# ── Variable tab column spec (shared across 5A-7) ─────────────────────────

VAR_HEADERS = [
    "Time Period", "Variable", "Label", "Values", "Definition",
    "Code Lists Group", "Additional Notes", "Date Modified\n(Initials + Date)",
    "QC Reviewed\n(Initials + Date)",
]
VAR_WIDTHS = [14, 16, 24, 18, 55, 16, 40, 18, 18]


def _write_variable_tab(ws, rows, section_title: str = ""):
    """Write a standardised variable tab with the shared column layout."""
    r = 1
    if section_title:
        _write_section_row(ws, r, section_title, len(VAR_HEADERS))
        r += 1

    # Instruction row
    ws.cell(row=r, column=1, value=(
        "Instruction: Delete any examples in grey italic text that are not "
        "relevant/applicable to the study once the programming specification has been finalized"
    ))
    r += 1

    # Column headers
    _write_header_row(ws, r, VAR_HEADERS)
    r += 1

    # Data rows
    for var_row in rows:
        values = [
            var_row.time_period,
            var_row.variable,
            var_row.label,
            var_row.values,
            var_row.definition,
            var_row.code_lists_group,
            var_row.additional_notes,
            var_row.date_modified,
            var_row.qc_reviewed,
        ]
        _write_data_row(ws, r, values, confidence=var_row.confidence)
        r += 1

    # Column widths
    for i, w in enumerate(VAR_WIDTHS):
        ws.column_dimensions[chr(65 + i)].width = w


def save_excel(spec: ProgramSpec, output_path: str) -> str:
    """Generate formatted Excel workbook from ProgramSpec (9-tab layout)."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required for Excel output. Run: pip install openpyxl")

    wb = Workbook()

    # ── Tab 1: Cover ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "1.Cover"

    # Section A: Study Information
    _write_section_row(ws, 1, "Section A: Study Information", 6)
    info_rows = [
        ("Study ID:", spec.study_info.study_id),
        ("Asset:", spec.study_info.asset),
        ("Indication:", spec.study_info.indication),
        ("Study Title:", spec.study_info.study_title),
        ("Study Status:", spec.study_info.study_status),
        ("Study Closed Date:", spec.study_info.study_closed_date),
    ]
    for r, (label, val) in enumerate(info_rows, 2):
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD
        ws.cell(row=r, column=2, value=val).font = FONT_NORMAL

    # Section B: Study Team
    team_start = len(info_rows) + 3
    _write_section_row(ws, team_start, "Section B: Study Team", 6)
    _write_header_row(ws, team_start + 1, ["Role", "Group", "Name"])
    for r, member in enumerate(spec.team_members, team_start + 2):
        ws.cell(row=r, column=1, value=member.role).font = FONT_NORMAL
        ws.cell(row=r, column=2, value=member.group).font = FONT_NORMAL
        ws.cell(row=r, column=3, value=member.name).font = FONT_NORMAL

    # Section C: Programming Specification Tabs
    tabs_start = team_start + len(spec.team_members) + 3
    _write_section_row(ws, tabs_start, "Section C: Programming Specification Tabs", 6)
    _write_header_row(ws, tabs_start + 1, ["Tab", "Purpose of Tab", "Status of Tab", "Notes", "Deadlines"])
    for r, ts in enumerate(spec.tab_statuses, tabs_start + 2):
        _write_data_row(ws, r, [ts.tab, ts.purpose, ts.status, ts.notes, ts.deadlines])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16

    # ── Tab 2: QC Review ───────────────────────────────────────────────────
    ws_qc = wb.create_sheet("2.QC Review")
    _write_header_row(ws_qc, 1, ["#", "Tab", "Comment", "Resolution", "Status"])
    for r, warning in enumerate(spec.qc_warnings, 2):
        _write_data_row(ws_qc, r, [r - 1, "", warning, "", "Open"])
    ws_qc.column_dimensions["A"].width = 6
    ws_qc.column_dimensions["B"].width = 16
    ws_qc.column_dimensions["C"].width = 60
    ws_qc.column_dimensions["D"].width = 40
    ws_qc.column_dimensions["E"].width = 12

    # ── Tab 3: Data Prep ───────────────────────────────────────────────────
    ws_dp = wb.create_sheet("3.Data Prep")
    r = 1

    # Section A: Data Source Selection
    _write_section_row(ws_dp, r, "Section A: Data Source Selection", 6)
    r += 1
    _write_header_row(ws_dp, r, ["Data source", "Population Subset", "Version"])
    r += 1
    _write_data_row(ws_dp, r, [
        spec.data_source.data_source,
        spec.data_source.population_subset,
        spec.data_source.version,
    ])
    r += 2

    # Section B: Define Important Dates
    _write_section_row(ws_dp, r, "Section B: Define Important Dates", 7)
    r += 1
    _write_header_row(ws_dp, r, [
        "Variable", "Label", "Date", "Definition", "Additional Notes",
        "Date Modified\n(Initials + Date)", "QC Reviewed\n(Initials + Date)",
    ])
    r += 1
    for dt in spec.important_dates:
        _write_data_row(ws_dp, r, [
            dt.variable, dt.label, dt.date_format, dt.definition,
            dt.additional_notes, dt.date_modified, dt.qc_reviewed,
        ], confidence=dt.confidence)
        r += 1
    r += 1

    # Section C: Define Study Time Periods
    _write_section_row(ws_dp, r, "Section C: Define Study Time Periods", 7)
    r += 1
    _write_header_row(ws_dp, r, [
        "Time Period", "Label", "Definition", "Additional Notes",
        "Date Modified\n(Initials + Date)", "QC Reviewed\n(Initials + Date)",
    ])
    r += 1
    for tp in spec.time_periods:
        _write_data_row(ws_dp, r, [
            tp.time_period, tp.label, tp.definition,
            tp.additional_notes, tp.date_modified, tp.qc_reviewed,
        ], confidence=tp.confidence)
        r += 1
    r += 1

    # Section D: Source Data Preparation
    _write_section_row(ws_dp, r, "Section D: Source Data Preparation", 7)
    r += 1
    _write_header_row(ws_dp, r, [
        "#", "Source Table/Variable", "Situation Requiring Resolution",
        "Action", "Reasoning / Comments",
        "Date Modified\n(Initials + Date)", "QC Reviewed\n(Initials + Date)",
    ])
    r += 1
    for sdp in spec.source_data_prep:
        _write_data_row(ws_dp, r, [
            sdp.row_number, sdp.source_table_variable, sdp.situation,
            sdp.action, sdp.reasoning, sdp.date_modified, sdp.qc_reviewed,
        ])
        r += 1

    ws_dp.column_dimensions["A"].width = 16
    ws_dp.column_dimensions["B"].width = 22
    ws_dp.column_dimensions["C"].width = 14
    ws_dp.column_dimensions["D"].width = 55
    ws_dp.column_dimensions["E"].width = 40
    ws_dp.column_dimensions["F"].width = 18
    ws_dp.column_dimensions["G"].width = 18

    # ── Tab 4: StudyPop ────────────────────────────────────────────────────
    ws_sp = wb.create_sheet("4.StudyPop")
    r = 1

    # Section A: Inclusion / Exclusion Criteria
    sp_headers = [
        "Time Period", "Variable", "Label", "Values", "Definition",
        "Additional Notes", "Code Lists Group",
        "Date Modified\n(Initials + Date)", "QC Reviewed\n(Initials + Date)",
    ]
    _write_section_row(ws_sp, r, "Section A: Inclusion / Exclusion Criteria", len(sp_headers))
    r += 1
    note = "Note to Programmer: Use the naming convention 'Time Period_Variable'"
    ws_sp.cell(row=r, column=1, value=note).font = FONT_BOLD
    r += 1
    _write_header_row(ws_sp, r, sp_headers)
    r += 1

    for crit in spec.inclusion_criteria:
        _write_data_row(ws_sp, r, [
            crit.time_period, crit.variable, crit.label, crit.values,
            crit.definition, crit.additional_notes, crit.code_lists_group,
            crit.date_modified, crit.qc_reviewed,
        ], confidence=crit.confidence)
        r += 1

    for crit in spec.exclusion_criteria:
        _write_data_row(ws_sp, r, [
            crit.time_period, crit.variable, crit.label, crit.values,
            crit.definition, crit.additional_notes, crit.code_lists_group,
            crit.date_modified, crit.qc_reviewed,
        ], confidence=crit.confidence)
        r += 1
    r += 1

    # Section B: Cohort Definitions
    _write_section_row(ws_sp, r, "Section B: Cohort Definition", len(sp_headers))
    r += 1
    _write_header_row(ws_sp, r, [
        "Variable", "Label", "Values", "Definition", "Additional Notes",
        "Code Lists Group", "Date Modified\n(Initials + Date)", "QC Reviewed\n(Initials + Date)",
    ])
    r += 1
    for coh in spec.cohort_definitions:
        _write_data_row(ws_sp, r, [
            coh.variable, coh.label, coh.values, coh.definition,
            coh.additional_notes, coh.code_lists_group,
            coh.date_modified, coh.qc_reviewed,
        ], confidence=coh.confidence)
        r += 1

    for i, w in enumerate([14, 16, 24, 18, 55, 40, 16, 18, 18]):
        ws_sp.column_dimensions[chr(65 + i)].width = w

    # ── Variable tabs (5A–7) ───────────────────────────────────────────────
    tab_configs = [
        ("5A.Demos", spec.demographics, "Section A: Variables for all patients"),
        ("5B.ClinChars", spec.clinical_characteristics, "Section A: Clinical Characteristics"),
        ("5C.BioVars", spec.biomarkers, "Section A: Biomarker Variables"),
        ("5D.LabVars", spec.lab_variables, "Section A: Laboratory Variables"),
        ("6.TreatVars", spec.treatment_variables, "Section A: Treatment Variables"),
        ("7.Outcomes", spec.outcome_variables, "Section A: Defining Analytic Variables"),
    ]

    for tab_name, rows, section_title in tab_configs:
        ws_tab = wb.create_sheet(tab_name)
        _write_variable_tab(ws_tab, rows, section_title=section_title)

    # Save
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)
